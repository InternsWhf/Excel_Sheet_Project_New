from flask import Flask, render_template, request, send_file
import os, base64, re, shutil, json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from dotenv import load_dotenv
import openai
from datetime import datetime

# Initialize Flask app
app = Flask(__name__)
load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")

@app.route('/')
def index():
    formats_path = "formats"
    os.makedirs(formats_path, exist_ok=True)

    # Only allow these templates
    allowed_templates = {"SHOT BLASTING.xlsx", "GRINDING.xlsx", "MPI.xlsx"}
    templates = [f for f in os.listdir(formats_path) if f in allowed_templates]
    return render_template("index.html", templates=templates)

@app.route('/upload', methods=['POST'])
def upload():
    image = request.files.get("image")
    template_name = request.form.get("format")

    if not image or not template_name:
        return "Missing image or template", 400

    # Convert image to base64
    image_b64 = base64.b64encode(image.read()).decode("utf-8")
    image_url = f"data:image/jpeg;base64,{image_b64}"

    # ================= Prompt Selection =================
    if "MPI" in template_name:
        prompt = """
You are an expert OCR model. Carefully extract structured table data from the attached MPI Production Book image.

Each row must include these fields:

- "Date"
- "Shift"
- "Machine No."
- "Operator Name"
- "Die No."
- "RF. NO"
- "Heat Code"
- "Head Shot"
- "Coil Shot"
- "Total Qty. Checked"
- "OK"
- "Rework"
- "Remark"

⚠️ If any field is blank or "-", preserve it exactly.
⚠️ Repeat metadata (Date, Shift, etc.) across all rows.
⚠️ Return only valid JSON array like:

[
  {
    "Date": "30/07/25", "Shift": "II", "Machine No.": "02", "Operator Name": "SHAMEBAZ",
    "Die No.": "4998", "RF. NO": "-", "Heat Code": "-", "Head Shot": "2500", "Coil Shot": "2800",
    "Total Qty. Checked": "81", "OK": "74", "Rework": "07", "Remark": "Fresh"
  },
  ...
]

Do not include any explanations — only JSON.
"""
    elif "SHOT BLASTING" in template_name or "GRINDING" in template_name:
        prompt = """
You are an expert OCR model. The image shows a handwritten table with **two sets of columns side-by-side**:

Left side: "Die No" and "Qty"  
Right side: "Die No" and "Qty" (may contain RSB or machining notes)

Extract ALL rows as they appear. Even if values are missing on one side, preserve empty cells.

JSON format:
[
  {"Die No": "5213", "Qty": "190", "Die No.1": "", "Qty.1": ""},
  {"Die No": "4209", "Qty": "169", "Die No.1": "RS:B", "Qty.1": ""},
  ...
]

⚠️ Return only valid JSON array — no extra text.
"""
    else:
        prompt = """
You are a document parser. Extract the entire table from the image, preserving all headers and row data.

The output should be a JSON array like:
[{"Date": "2024-07-01", "Code": "A123", "Heat Code": "...", ...}]

⚠️ All column names must match exactly with the Excel template.
⚠️ Return only JSON. No comments.
"""

    # ============== GPT-4o API Call ==============
    try:
        response = openai.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": prompt.strip()},
                {"role": "user", "content": [{"type": "image_url", "image_url": {"url": image_url}}]},
            ],
            temperature=0
        )

        result = response.choices[0].message.content

        # Extract valid JSON block
        match = re.search(r"\[\s*{.*?}\s*\]", result, re.DOTALL)
        if not match:
            return "❌ No valid JSON found in GPT response", 500

        json_str = match.group(0)

        # Save for debugging
        with open("debug_output.json", "w", encoding="utf-8") as f:
            f.write(json_str)

        df = pd.read_json(BytesIO(json_str.encode()))
        print("📊 Parsed DataFrame:\n", df.head())

    except Exception as e:
        return f"❌ GPT-4o failed: {e}", 500

    # ============== Excel File Fill ==============
    output_dir = "outputs"
    os.makedirs(output_dir, exist_ok=True)

    template_path = os.path.join("formats", template_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"filled_output_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_file)
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Write headers in Row 2 (skip merged)
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=2, column=j+1)
        if not isinstance(cell, MergedCell):
            cell.value = col

    # Write data from Row 3 onward
    for i, row in df.iterrows():
        for j, val in enumerate(row):
            cell = ws.cell(row=i+3, column=j+1)
            if not isinstance(cell, MergedCell):
                cell.value = val

    wb.save(output_path)
    print(f"✅ Excel saved: {output_path}")
    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    print("✅ Flask OCR server started at http://127.0.0.1:5000")
    app.run(debug=True)
