from flask import Flask, render_template, request, send_file
import os, base64, re, shutil, json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from dotenv import load_dotenv
from datetime import datetime
from PIL import Image
from pdf2image import convert_from_bytes
import openai

# Initialize Flask App
app = Flask(__name__)
load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")


@app.route('/')
def index():
    formats_path = "formats"
    os.makedirs(formats_path, exist_ok=True)
    templates = os.listdir(formats_path)
    return render_template("index.html", templates=templates)


@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get("image")
    template_name = request.form.get("format")

    if not file or not template_name:
        return "Missing image or template", 400

    filename = file.filename.lower()

    try:
        # ✅ Handle PDF Uploads
        if filename.endswith(".pdf"):
            pages = convert_from_bytes(file.read(), dpi=300, first_page=1, last_page=1)
            if not pages:
                return "❌ No page found in PDF", 400
            image_pil = pages[0]
        else:
            image_pil = Image.open(file.stream)

        # ✅ Convert image to base64
        buffered = BytesIO()
        image_pil.save(buffered, format="JPEG")
        image_b64 = base64.b64encode(buffered.getvalue()).decode("utf-8")
        image_url = f"data:image/jpeg;base64,{image_b64}"

    except Exception as e:
        return f"❌ Image or PDF processing failed: {e}", 500

    # ========== PROMPT SELECTION ==========

    if "GRINDING" in template_name.upper():
        prompt = """
You are an expert OCR model. Carefully extract all rows from the handwritten DAILY GRINDING REPORT form.

Each row includes:
- DATE
- SHIFT
- DIE NO
- NET WT.
- GRINDING QTY
- STATUS
- VENDOR

⚠️ Extract every row exactly as it appears — from top to bottom, left to right.
⚠️ If any field contains multiple parts (e.g. "OD", "(4462)"), preserve all parts, comma-separated.
⚠️ Use "-" where dash is written and retain special symbols or spacing in values like "443-20", "PLW", "SAARAMBHA".

🧾 Output format must be a JSON array like:
[
  {
    "DATE": "25/07/25",
    "SHIFT": "I",
    "DIE NO": "5196",
    "NET WT.": "250",
    "GRINDING QTY": "",
    "STATUS": "",
    "VENDOR": ""
  },
  ...
]

⚠️ All rows must include date and shift.
⚠️ Do not summarize, skip, or comment. Return ONLY the JSON array.
"""
    else:
        prompt = """
You are an expert OCR model. The image shows a handwritten table with **two sets of columns side-by-side**:

- Left side: "Die No" and "Qty"
- Right side: "Die No" and "Qty" (may contain RSB notes, machining remarks, or be blank)

🔍 Your task:
- Extract **ALL rows** as they appear, row by row.
- Even if values are missing on one side, preserve empty cells to match visual structure.

✅ Output format must be a JSON array like:
[
  {"Die No": "5213", "Qty": "190", "Die No.1": "", "Qty.1": ""},
  {"Die No": "4209", "Qty": "169", "Die No.1": "RS:B", "Qty.1": ""},
  {"Die No": "5213", "Qty": "20", "Die No.1": "RS:B Machining", "Qty.1": ""},
  {"Die No": "", "Qty": "", "Die No.1": "(हाथ से)", "Qty.1": "Machine"}
]

⚠️ Don't skip rows or join cells. Empty entries must be left blank (e.g., "").
⚠️ Return ONLY the valid JSON array — no explanation or comments.
"""

    # ✅ Call GPT-4o Vision API
    try:
        response = openai.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": prompt.strip()},
                {"role": "user", "content": [{"type": "image_url", "image_url": {"url": image_url}}]},
            ],
            temperature=0
        )

        result = response.choices[0].message.content
        match = re.search(r"\[\s*{.*?}\s*\]", result, re.DOTALL)
        if not match:
            return "❌ No valid JSON found in GPT response", 500

        json_str = match.group(0)

        with open("debug_output.json", "w", encoding="utf-8") as debug_file:
            debug_file.write(json_str)

        df = pd.read_json(BytesIO(json_str.encode()))
        print("📊 Parsed DataFrame:\n", df.head())

    except Exception as e:
        return f"❌ GPT-4o failed: {e}", 500

    # ✅ Prepare Excel output
    output_dir = "outputs"
    os.makedirs(output_dir, exist_ok=True)

    template_path = os.path.join("formats", template_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"filled_output_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_file)
    shutil.copy(template_path, output_path)

    try:
        wb = load_workbook(output_path)
        ws = wb.active

        for j, col in enumerate(df.columns):
            cell = ws.cell(row=2, column=j+1)
            if not isinstance(cell, MergedCell):
                cell.value = col

        for i, row in df.iterrows():
            for j, val in enumerate(row):
                cell = ws.cell(row=i+3, column=j+1)
                if not isinstance(cell, MergedCell):
                    cell.value = val

        wb.save(output_path)
        print(f"✅ Excel saved: {output_path}")

    except Exception as e:
        return f"❌ Excel Write failed: {e}", 500

    return send_file(output_path, as_attachment=True)


if __name__ == '__main__':
    print("✅ Flask OCR server started at http://127.0.0.1:5000")
    app.run(debug=True)
